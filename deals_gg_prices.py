"""This module get game prices on deals.gg"""
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook

FILE = 'Pricelist.xlsx'


def get_price(gamename):
    "return deals.gg price for a game game"
    game = gamename.rstrip().replace(" ", "+")
    url = "https://gg.deals/games/?title=" + game
    res = requests.get(url)
    soup = bs(res.content, 'html.parser')
    price = soup.find("span", class_='numeric')
    if not price:
        price = "Not found"
        name = "Not found"
        print(gamename + " => Not found")
    else:
        fulllink = "https://gg.deals" + \
            price.find_parent(
                "div", class_='game-list-item').select_one('.full-link')['href']
        req = requests.get(fulllink)
        soup = bs(req.content, 'html.parser')
        price = soup.select_one('.best-deal')
        if not price:
            price = soup.select_one('.lowest-recorded.price-hl.price-widget')
        if not price:
            price = soup.select_one('.numeric').text.replace("~", "")
        else:
            price = price.select_one('.numeric').text.replace("~", "")
        name = soup.select_one('h1').text[4:]
        print(gamename + " => " + name + ": " + price)

    return price, name


def main():
    """Main func"""
    choice = 0
    while choice < 1 or choice > 2:
        choice = int(
            input("Enter :\n1 to enter a game name\n2 to update the excel file\n"))

    if choice == 1:
        gamename = input("Enter the game name :\n")
        get_price(str(gamename))
    else:
        wbk = load_workbook(filename=FILE)
        wsh = wbk.active
        print("\n\n------------------\nSEARCHING PRICES\n------------------\n")
        for i in range(3, wsh.max_row):
            name = wsh['A'+str(i)].value
            if name:
                print(name)
                wsh['E' + str(i)], wsh['F' + str(i)] = get_price(str(name))

        wbk.save(FILE)
        print("\n\nDONE!\nCheck Pricelist.xlsx to get the prices")

    input("Press any key to quit...")


if __name__ == "__main__":
    main()
