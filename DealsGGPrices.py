import sys
import requests
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook
from openpyxl import load_workbook
from os import system, name

file = 'Pricelist.xlsx'

def getPrice( gamename ):
    "return deals.gg price for a given game"
    game = gamename.rstrip().replace(" ", "+")
    url = "https://gg.deals/games/?title=" + game
    r = requests.get(url)
    soup = bs(r.content, 'html.parser')
    price = soup.find("span", class_='numeric')
    if not price:
        price = "Not found"
        name = "Not found"
        print(gamename + " => Not found")
    else:
        fulllink = "https://gg.deals" + price.find_parent("div", class_='game-list-item').select_one('.full-link')['href']
        req = requests.get(fulllink)
        soup = bs(req.content, 'html.parser')
        price = soup.select_one('.best-deal')
        if not price:
            price = soup.select_one('.lowest-recorded.price-hl.price-widget')
        if not price:
            price = soup.select_one('.numeric').text.replace("~", "")
        else:
            price = price.select_one('.numeric').text.replace("~", "")
        name = soup.select_one('h1').text[4:]
        print(gamename + " => " + name + ": " + price)
        
    return price, name
    
c = 0
while c < 1 or c > 2:
    c = int(input("Enter :\n1 to enter a game name\n2 to update the excel file\n"))

if c == 1:
    gamename = input("Enter the game name :\n")
    getPrice(str(gamename))
else:
    wb = load_workbook(filename = file)
    ws = wb.active
    i = 2
    print("\n\n------------------\nSEARCHING PRICES\n------------------\n")
    for row in ws.iter_rows(min_row=1, max_col=5):
        i += 1
        if i > len(list(ws.rows)):
            break
        else:
            name = ws['A'+str(i)].value
            if name:
                ws['E'+str(i)], ws['F'+str(i)] = getPrice(str(name))
    wb.save(file)
    print("\n\nDONE!\nCheck Pricelist.xlsx to get the prices")

input("Press any key to quit...")